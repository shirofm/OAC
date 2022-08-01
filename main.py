import openpyxl
import time
from selenium.webdriver.support import expected_conditions as EC
# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.wait import WebDriverWait


def consultaSC(contents, animal):
    ws = wb[animal]
    mes = int(contents[0])
    linha = 25
    for i in range(int(contents[0])):
        ws[f'c{linha}'] = driver.find_element(By.XPATH,
                                              f'//*[@id="app-table-responsive"]/table/tbody/tr[{mes}]/td[3]/div').text.replace(
            '$', '').replace('.', '')
        ws[f'd{linha}'] = driver.find_element(By.XPATH,
                                              f'//*[@id="app-table-responsive"]/table/tbody/tr[{mes}]/td[4]/div').text.replace(
            '.', '')
        linha += 1
        mes -= 1
def consultaBR(contents, animal):
    ws = wb[animal]
    mes = int(contents[0])
    linha = 104
    for i in range(int(contents[0])):
        ws[f'c{linha}'] = driver.find_element(By.XPATH,
                                              f'//*[@id="app-table-responsive"]/table/tbody/tr[{mes}]/td[4]/div').text.replace(
            '$', '').replace('.', '')
        ws[f'd{linha}'] = driver.find_element(By.XPATH,
                                              f'//*[@id="app-table-responsive"]/table/tbody/tr[{mes}]/td[5]/div').text.replace(
            '.', '')
        linha += 1
        mes -= 1

def consulta(contents):
    for i in range (1, 11):
        driver.get(contents[i])
        time.sleep(4)
        click = driver.find_element(By.XPATH, "//*[@id='form']/div[10]/button[2]")
        click.click()
        if i == 1:
            consultaSC(contents, 'FRANGOS')
            print('Atualizando Frangos SC..')
        if i == 2:
            consultaBR(contents, 'FRANGOS')
            print('Atualizando Frangos BR..')
        if i == 3:
            consultaSC(contents, 'SUÍNOS')
            print('Atualizando Suinos SC..')
        if i == 4:
            consultaBR(contents, 'SUÍNOS')
            print('Atualizando Suinos BR..')
        if i == 5:
            consultaSC(contents, 'BOVINOS')
            print('Atualizando Bovinos SC..')
        if i == 6:
            consultaBR(contents, 'BOVINOS')
            print('Atualizando Bovinos BR..')
        if i == 7:
            consultaSC(contents, 'PERUS')
            print('Atualizando Perus SC..')
        if i == 8:
            consultaBR(contents, 'PERUS')
            print('Atualizando Perus BR..')
        if i == 9:
            consultaSC(contents, 'PATOS E MARRECOS')
            print('Atualizando Patos SC..')
        if i == 10:
            consultaBR(contents, 'PATOS E MARRECOS')
            print('Atualizando Patos BR..')



if __name__ == '__main__':
    file = 'export.xlsx'
    wb = load_workbook(file)
    PATH = 'chromedriver.exe'
    service = Service('chromedriver.exe')
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(10)
    fileToRead = 'links.txt'
    with open(fileToRead) as f:
        content = f.readlines()
        content = [x.strip() for x in content if not x.startswith('#')]
    print(content)
    consulta(content)
    #ws = wb['FRANGOS']
    #ws['C31'] = 'teste'
    wb.save(file)
    print('Terminando programa..')
    driver.quit()
